import streamlit as st
import os
import sqlite3
import time
import base64
import io
import openpyxl
from openpyxl.styles import Font
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(BASE_DIR, "assets")
DB_PATH = os.path.join(BASE_DIR, "client_leads.db")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")

logo_path = os.path.join(ASSETS_DIR, "logo.png")
favicon = logo_path if os.path.exists(logo_path) else "⚡"

st.set_page_config(page_title="Nexus Digital", page_icon=favicon, layout="wide")

def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS client_inquiries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL,
            phone_number TEXT NOT NULL,
            service_required TEXT NOT NULL,
            requirements TEXT NOT NULL,
            submitted_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            status TEXT DEFAULT 'New'
        )
    ''')
    conn.commit()
    conn.close()

init_db()

def load_10_10_css():
    st.markdown("""
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <link href="https://fonts.googleapis.com/css2?family=Fira+Code:wght@400;600&family=Inter:wght@400;600;800&display=swap" rel="stylesheet">
    <style>
        :root {
            --bg-color: #030303;
            --accent-primary: #10B981;
            --accent-secondary: #059669;
            --accent-tertiary: #06B6D4;
            --text-main: #F8FAFC;
            --text-muted: #9CA3AF;
            --card-bg: rgba(255, 255, 255, 0.02);
            --card-border: rgba(255, 255, 255, 0.08);
        }
        
        /* Ambient Background Orbs */
        body::before {
            content: '';
            position: fixed;
            top: -10%;
            left: -10%;
            width: 50vw;
            height: 50vw;
            background: radial-gradient(circle, rgba(16,185,129,0.1) 0%, rgba(0,0,0,0) 70%);
            filter: blur(120px);
            z-index: -1;
            pointer-events: none;
        }
        body::after {
            content: '';
            position: fixed;
            bottom: -10%;
            right: -10%;
            width: 60vw;
            height: 60vw;
            background: radial-gradient(circle, rgba(6,182,212,0.08) 0%, rgba(0,0,0,0) 70%);
            filter: blur(150px);
            z-index: -1;
            pointer-events: none;
        }

        /* Scrollbar & Selection */
        ::-webkit-scrollbar {
            width: 10px;
        }
        ::-webkit-scrollbar-track {
            background: #030303;
        }
        ::-webkit-scrollbar-thumb {
            background: #333;
            border-radius: 5px;
        }
        ::-webkit-scrollbar-thumb:hover {
            background: var(--accent-primary);
        }
        ::selection {
            background: rgba(16, 185, 129, 0.3);
            color: #fff;
        }

        .stApp {
            background-color: transparent !important;
            color: var(--text-main);
            font-family: 'Inter', sans-serif;
        }
        
        /* Animations */
        @keyframes fadeInUp {
            from { opacity: 0; transform: translateY(30px); }
            to { opacity: 1; transform: translateY(0); }
        }
        .animate-up {
            animation: fadeInUp 0.8s cubic-bezier(0.16, 1, 0.3, 1) forwards;
            opacity: 0;
        }
        .delay-1 { animation-delay: 0.1s; }
        .delay-2 { animation-delay: 0.2s; }
        .delay-3 { animation-delay: 0.3s; }
        .delay-4 { animation-delay: 0.4s; }

        @keyframes floatAnim {
            0% { transform: translateY(0px); }
            50% { transform: translateY(-8px); }
            100% { transform: translateY(0px); }
        }
        .float-hover {
            animation: floatAnim 6s ease-in-out infinite;
            transition: all 0.4s cubic-bezier(0.16, 1, 0.3, 1);
        }
        .float-hover:hover {
            animation-play-state: paused;
            transform: translateY(-10px) scale(1.02) !important;
            box-shadow: 0 15px 35px rgba(16, 185, 129, 0.2);
            border-color: var(--accent-primary);
        }

        h1, h2, h3, h4, h5, h6 {
            color: var(--text-main) !important;
            font-family: 'Inter', sans-serif;
            font-weight: 800;
        }
        
        .hero-headline {
            font-size: 4.5rem;
            font-weight: 800;
            color: var(--text-main);
            margin-bottom: 1rem;
            text-align: center;
            line-height: 1.1;
            letter-spacing: -2px;
        }
        .hero-headline span {
            background: linear-gradient(135deg, var(--accent-primary), var(--accent-tertiary));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }
        .hero-subhead {
            font-size: 1.25rem;
            color: var(--text-muted);
            margin-bottom: 3rem;
            text-align: center;
            max-width: 800px;
            margin-left: auto;
            margin-right: auto;
            line-height: 1.6;
        }
        
        /* Glassmorphism Bento Card Styles */
        .bento-card {
            background-color: var(--card-bg);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border: 1px solid var(--card-border);
            border-radius: 20px;
            padding: 32px;
            margin-bottom: 20px;
            transition: all 0.4s cubic-bezier(0.16, 1, 0.3, 1);
            height: 100%;
            display: flex;
            flex-direction: column;
            position: relative;
            overflow: hidden;
        }
        .bento-card::before {
            content: '';
            position: absolute;
            top: 0; left: 0; width: 100%; height: 100%;
            background: linear-gradient(135deg, rgba(255,255,255,0.05) 0%, rgba(255,255,255,0) 100%);
            opacity: 0;
            transition: opacity 0.4s ease;
            pointer-events: none;
        }
        .bento-card:hover {
            border-color: rgba(16, 185, 129, 0.4);
            transform: translateY(-5px);
            box-shadow: 0 20px 40px rgba(0, 0, 0, 0.4), 0 0 20px rgba(16, 185, 129, 0.1);
        }
        .bento-card:hover::before {
            opacity: 1;
        }
        .bento-card h3 {
            color: var(--text-main) !important;
            margin-top: 0;
            font-size: 1.75rem;
            margin-bottom: 12px;
            letter-spacing: -0.5px;
        }
        .bento-card p {
            color: var(--text-muted);
            flex-grow: 1;
            margin-bottom: 20px;
            font-size: 1rem;
            line-height: 1.6;
        }
        
        .gradient-stat {
            font-size: 4.5rem;
            font-weight: 800;
            margin: 0;
            background: linear-gradient(135deg, var(--accent-primary), var(--accent-tertiary));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            line-height: 1;
        }
        
        /* Monospace Accents */
        .mono-text {
            font-family: 'Fira Code', monospace;
            font-size: 0.85rem;
            color: var(--accent-primary);
            letter-spacing: 1px;
            text-transform: uppercase;
            margin-bottom: 15px;
        }
        
        .pricing-tier {
            background: rgba(0,0,0,0.5);
            border-radius: 12px;
            padding: 16px;
            margin-top: 15px;
            border: 1px solid rgba(255,255,255,0.05);
            font-family: 'Fira Code', monospace;
            font-size: 0.85rem;
        }
        .pricing-tier strong {
            color: var(--accent-primary);
            font-family: 'Inter', sans-serif;
            font-size: 1rem;
            display: inline-block;
            margin-bottom: 4px;
        }
        
        /* Buttons */
        .stButton>button {
            background-color: transparent !important;
            color: var(--text-main) !important;
            border: 1px solid rgba(255, 255, 255, 0.15) !important;
            border-radius: 12px !important;
            font-family: 'Inter', sans-serif;
            font-weight: 600;
            transition: all 0.3s cubic-bezier(0.16, 1, 0.3, 1);
            width: 100%;
            padding: 12px 24px !important;
        }
        .stButton>button:hover {
            border-color: var(--accent-primary) !important;
            color: var(--accent-primary) !important;
            background-color: rgba(16, 185, 129, 0.05) !important;
        }
        
        /* Primary CTA Button */
        button[kind="primary"] {
            background: linear-gradient(135deg, var(--accent-primary), var(--accent-secondary)) !important;
            color: #000000 !important;
            border: none !important;
            border-radius: 12px !important;
            font-weight: 800;
            font-size: 1.1rem;
            padding: 15px 30px !important;
            box-shadow: 0 4px 15px rgba(16, 185, 129, 0.2);
            transition: all 0.3s ease;
        }
        button[kind="primary"]:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 25px rgba(16, 185, 129, 0.4);
            color: #ffffff !important;
        }
        
        /* Form Inputs */
        div[data-baseweb="input"] > div, div[data-baseweb="textarea"] > div, div[data-baseweb="select"] > div {
            background-color: rgba(255, 255, 255, 0.03);
            border: 1px solid rgba(255, 255, 255, 0.1);
            border-radius: 10px;
            color: var(--text-main);
            transition: all 0.3s ease;
        }
        div[data-baseweb="input"] > div:focus-within, div[data-baseweb="textarea"] > div:focus-within {
            border-color: var(--accent-primary);
            background-color: rgba(16, 185, 129, 0.05);
        }
        
        hr { border-color: rgba(255, 255, 255, 0.05); opacity: 1; margin: 6rem 0; }
        
        /* Social Proof Banner */
        .social-proof {
            display: flex;
            justify-content: center;
            align-items: center;
            flex-wrap: wrap;
            gap: 60px;
            padding: 30px 0;
            border-top: 1px solid rgba(255, 255, 255, 0.05);
            border-bottom: 1px solid rgba(255, 255, 255, 0.05);
            margin-bottom: 5rem;
            color: var(--text-muted);
            font-size: 0.95rem;
            font-weight: 600;
            letter-spacing: 2px;
            text-transform: uppercase;
            font-family: 'Fira Code', monospace;
        }
        .social-proof span {
            display: flex;
            align-items: center;
            gap: 12px;
            opacity: 0.7;
            transition: opacity 0.3s ease;
        }
        .social-proof span:hover {
            opacity: 1;
            color: var(--accent-primary);
        }
        .social-proof i {
            font-size: 1.5rem;
        }
        
        /* Process Steps */
        .step-card {
            text-align: center;
            padding: 20px;
            transition: transform 0.3s ease;
        }
        .step-card:hover {
            transform: translateY(-5px);
        }
        .step-number {
            background: rgba(16, 185, 129, 0.1);
            color: var(--accent-primary);
            width: 50px;
            height: 50px;
            border-radius: 12px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-family: 'Fira Code', monospace;
            font-weight: 800;
            font-size: 1.2rem;
            margin: 0 auto 20px auto;
            border: 1px solid rgba(16, 185, 129, 0.3);
        }
        .step-title {
            font-weight: 800;
            color: var(--text-main);
            margin-bottom: 12px;
            font-size: 1.25rem;
        }
        .step-desc {
            color: var(--text-muted);
            font-size: 1rem;
            line-height: 1.6;
        }

        /* Sidebar Styling */
        [data-testid="stSidebar"] {
            background-color: rgba(3, 3, 3, 0.8);
            backdrop-filter: blur(20px);
            border-right: 1px solid rgba(255, 255, 255, 0.05);
        }
        [data-testid="stSidebarNav"] {
            display: none;
        }
        
        /* Section Headers */
        .section-title {
            text-align: center;
            font-size: 3rem;
            font-weight: 800;
            margin-bottom: 0.5rem;
            letter-spacing: -1px;
            color: var(--text-main);
        }
        .section-subtitle {
            text-align: center;
            color: var(--text-muted);
            margin-bottom: 4rem;
            font-size: 1.2rem;
            max-width: 600px;
            margin-left: auto;
            margin-right: auto;
            line-height: 1.6;
        }

        /* Footer */
        .footer {
            text-align: center;
            padding: 40px 0;
            margin-top: 60px;
            border-top: 1px solid rgba(255,255,255,0.05);
            font-family: 'Fira Code', monospace;
            font-size: 0.85rem;
            color: var(--text-muted);
            opacity: 0.6;
        }

        /* Mobile Responsiveness */
        @media (max-width: 768px) {
            .hero-headline { font-size: 2.5rem; letter-spacing: -1px; }
            .hero-subhead { font-size: 1rem; margin-bottom: 2rem; }
            .gradient-stat { font-size: 3.5rem; }
            .bento-card { padding: 20px; margin-bottom: 15px; }
            .social-proof { gap: 15px; font-size: 0.75rem; flex-wrap: wrap; justify-content: center; padding: 20px 0; }
            .social-proof i { font-size: 1.2rem; }
            .section-title { font-size: 2rem; }
            .section-subtitle { font-size: 1rem; margin-bottom: 2rem; }
            .footer { padding: 20px 0; font-size: 0.75rem; }
        }
    </style>
    """, unsafe_allow_html=True)

load_10_10_css()

def get_base64_of_bin_file(bin_file):
    if os.path.exists(bin_file):
        with open(bin_file, 'rb') as f:
            return base64.b64encode(f.read()).decode()
    return ""

if "page" not in st.session_state:
    st.session_state.page = "Main App"
if "show_form" not in st.session_state:
    st.session_state.show_form = False

def navigate(page):
    st.session_state.page = page

with st.sidebar:
    st.markdown("### Navigation")
    if st.button("Main App"): navigate("Main App")
    st.markdown("---")
    if st.button("Admin Panel"): navigate("Admin")

# =========================================================
# MAIN APP (Single Page)
# =========================================================
if st.session_state.page == "Main App":
    st.markdown("<br>", unsafe_allow_html=True)
    
    if os.path.exists(logo_path):
        logo_b64 = get_base64_of_bin_file(logo_path)
        st.markdown(f'''
        <div class="animate-up" style="display: flex; justify-content: center; margin-bottom: 20px; position: relative;">
            <img src="data:image/png;base64,{logo_b64}" style="height: 180px; width: auto; object-fit: contain; mix-blend-mode: screen; opacity: 0.9; filter: drop-shadow(0 0 30px rgba(16,185,129,0.15));">
        </div>
        ''', unsafe_allow_html=True)
            
    st.markdown('<div class="hero-headline animate-up">Talent Doesn\'t Open Doors.<br><span>Presentation Does.</span></div>', unsafe_allow_html=True)
    st.markdown('<div class="hero-subhead animate-up delay-1">You don\'t lose opportunities because you lack skill—you lose them because your presentation lets you down. We help tech students, graduates, and professionals stand out with elite portfolios and ATS resumes.</div>', unsafe_allow_html=True)
    
    st.markdown("""
    <div class="social-proof animate-up delay-2">
        <span><i class="fa-brands fa-python"></i> Python</span>
        <span><i class="fa-solid fa-code"></i> Streamlit</span>
        <span><i class="fa-solid fa-database"></i> SQL</span>
        <span><i class="fa-solid fa-robot"></i> AI SDKs</span>
        <span><i class="fa-brands fa-react"></i> React</span>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown('<div class="section-title animate-up delay-3">How It Works</div>', unsafe_allow_html=True)
    step1, step2, step3 = st.columns(3)
    with step1:
        st.markdown("""
        <div class="step-card animate-up delay-1">
            <div class="step-number">01</div>
            <div class="step-title">Consultation</div>
            <div class="step-desc">We align on your career goals, target roles, and specific technical requirements.</div>
        </div>
        """, unsafe_allow_html=True)
    with step2:
        st.markdown("""
        <div class="step-card animate-up delay-2">
            <div class="step-number">02</div>
            <div class="step-title">Development</div>
            <div class="step-desc">I engineer a custom, high-performance solution with premium UI/UX.</div>
        </div>
        """, unsafe_allow_html=True)
    with step3:
        st.markdown("""
        <div class="step-card animate-up delay-3">
            <div class="step-number">03</div>
            <div class="step-title">Deployment</div>
            <div class="step-desc">You launch with confidence and stand out instantly to top-tier recruiters.</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="section-title">Services & Pricing</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-subtitle">Transparent pricing. Uncompromising quality.</div>', unsafe_allow_html=True)
    
    # Bento Grid Layout
    # Row 1
    r1c1, r1c2 = st.columns([2, 1])
    with r1c1:
        st.markdown("""
        <div class="bento-card">
            <div class="mono-text">Service_01</div>
            <h3>Portfolio Websites</h3>
            <p>Custom-built, ultra-responsive, and interactive portfolios designed to showcase your unique value proposition. Powered by modern frameworks and deployed seamlessly.</p>
            <div class="pricing-tier">
                <strong>Standard: ₹1000</strong> - Multi-page, SEO, contact form.<br>
                <strong>Premium: ₹2500</strong> - Next.js/React, animations, DB integration.
            </div>
        </div>
        """, unsafe_allow_html=True)
        
    with r1c2:
        st.markdown("""
        <div class="bento-card" style="align-items: center; justify-content: center; text-align: center;">
            <div class="gradient-stat">50+</div>
            <p style="font-weight: 800; color: var(--text-main); margin-top: 10px; margin-bottom: 5px; font-size: 1.2rem;">Projects Delivered</p>
            <div class="mono-text" style="opacity: 0.8; margin-bottom: 0;">98% SATISFACTION</div>
            <div class="mono-text" style="opacity: 0.8;">48h TURNAROUND</div>
        </div>
        """, unsafe_allow_html=True)

    # Row 2
    r2c1, r2c2 = st.columns([1, 2])
    with r2c1:
        st.markdown("""
        <div class="bento-card">
            <div class="mono-text">Service_02</div>
            <h3>Resume Building</h3>
            <p>Engineered specifically for ATS parsers to ensure you bypass algorithmic filters and land the interview.</p>
            <div class="pricing-tier">
                <strong>Standard: ₹700</strong><br>
                <strong>Premium: ₹1500</strong>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
    with r2c2:
        st.markdown("""
        <div class="bento-card">
            <div class="mono-text">Service_03</div>
            <h3>Project Development</h3>
            <p>End-to-end Python automation, intelligent web scraping pipelines, and complex AI SDK integrations.</p>
            <div class="pricing-tier">
                <strong>Tech Stack:</strong> Python, Node.js, Gemini API, OpenAI, Selenium
            </div>
            <div class="pricing-tier" style="margin-top: 8px;">
                <strong>Custom Pricing</strong> - Based on scope and API requirements.
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="section-title">Featured Work</div>', unsafe_allow_html=True)
    
    # CSS for Slideshows & Resume Scores
    st.markdown("""
    <style>
    .slider {
        width: 100%;
        overflow: hidden;
        border-radius: 6px;
        position: relative;
    }
    .slides-3 {
        display: flex;
        width: 300%;
        animation: slide3 12s infinite cubic-bezier(0.8, 0, 0.2, 1);
    }
    .slides-3 img {
        width: 33.333%;
        object-fit: cover;
        border: 1px solid rgba(255,255,255,0.05);
    }
    @keyframes slide3 {
        0%, 25% { transform: translateX(0); }
        33.33%, 58.33% { transform: translateX(-33.333%); }
        66.66%, 91.66% { transform: translateX(-66.666%); }
        100% { transform: translateX(0); }
    }
    
    .slides-4 {
        display: flex;
        width: 400%;
        animation: slide4 16s infinite cubic-bezier(0.8, 0, 0.2, 1);
    }
    .slides-4 img {
        width: 25%;
        object-fit: cover;
        border: 1px solid rgba(255,255,255,0.05);
    }
    @keyframes slide4 {
        0%, 20% { transform: translateX(0); }
        25%, 45% { transform: translateX(-25%); }
        50%, 70% { transform: translateX(-50%); }
        75%, 95% { transform: translateX(-75%); }
        100% { transform: translateX(0); }
    }
    
    </style>
    """, unsafe_allow_html=True)
    
    col_port, col_ai = st.columns(2)
    with col_port:
        st.markdown("<h4 style='text-align:center;'>Portfolio Design Showcase</h4>", unsafe_allow_html=True)
        v1 = get_base64_of_bin_file(os.path.join(ASSETS_DIR, "veekshitha_1.png"))
        v2 = get_base64_of_bin_file(os.path.join(ASSETS_DIR, "veekshitha_2.png"))
        v3 = get_base64_of_bin_file(os.path.join(ASSETS_DIR, "veekshitha_3.png"))
        
        st.markdown(f'''
        <div class="bento-card" style="padding: 10px; border-radius: 12px; margin-bottom: 10px;">
            <div class="slider">
                <div class="slides-3">
                    <img src="data:image/png;base64,{v1}">
                    <img src="data:image/png;base64,{v2}">
                    <img src="data:image/png;base64,{v3}">
                </div>
            </div>
        </div>
        ''', unsafe_allow_html=True)
        
        with st.expander("🔍 Enlarge Gallery"):
            st.image(os.path.join(ASSETS_DIR, "veekshitha_1.png"), use_container_width=True)
            st.image(os.path.join(ASSETS_DIR, "veekshitha_2.png"), use_container_width=True)
            st.image(os.path.join(ASSETS_DIR, "veekshitha_3.png"), use_container_width=True)

    with col_ai:
        st.markdown("<h4 style='text-align:center;'>AI Assistant Pipeline Showcase</h4>", unsafe_allow_html=True)
        h1 = get_base64_of_bin_file(os.path.join(ASSETS_DIR, "health_ai_1.png"))
        h2 = get_base64_of_bin_file(os.path.join(ASSETS_DIR, "health_ai_2.png"))
        h3 = get_base64_of_bin_file(os.path.join(ASSETS_DIR, "health_ai_3.png"))
        h4 = get_base64_of_bin_file(os.path.join(ASSETS_DIR, "health_ai_4.png"))
        
        st.markdown(f'''
        <div class="bento-card" style="padding: 10px; border-radius: 12px; margin-bottom: 10px;">
            <div class="slider">
                <div class="slides-4">
                    <img src="data:image/png;base64,{h1}">
                    <img src="data:image/png;base64,{h2}">
                    <img src="data:image/png;base64,{h3}">
                    <img src="data:image/png;base64,{h4}">
                </div>
            </div>
        </div>
        ''', unsafe_allow_html=True)
        
        with st.expander("🔍 Enlarge Gallery"):
            st.image(os.path.join(ASSETS_DIR, "health_ai_1.png"), use_container_width=True)
            st.image(os.path.join(ASSETS_DIR, "health_ai_2.png"), use_container_width=True)
            st.image(os.path.join(ASSETS_DIR, "health_ai_3.png"), use_container_width=True)
            st.image(os.path.join(ASSETS_DIR, "health_ai_4.png"), use_container_width=True)

    # Resume section with animations via CSS + JavaScript injection
    resume_data = [
        {"file": "Sayyed_Raheem_Resume.jpg", "name": "The Executive", "score": 96},
        {"file": "Sayyed_Raheem_Resume_2.jpg", "name": "The Technologist", "score": 92},
        {"file": "Sayyed_Raheem_Resume_3.jpg", "name": "The Minimalist", "score": 98},
        {"file": "resume 4_page-0001.jpg", "name": "The Classic Professional", "score": 99},
    ]

    # Inject CSS animations + JS that will target resume columns
    st.markdown("""
    <style>
        @keyframes resEnter {
            0% { opacity: 0; transform: translateY(40px) scale(0.95); }
            100% { opacity: 1; transform: translateY(0) scale(1); }
        }
        @keyframes resFloat {
            0%, 100% { transform: translateY(0px); }
            50% { transform: translateY(-8px); }
        }
        @keyframes atsGlow {
            0%, 100% { box-shadow: 0 4px 15px rgba(16, 185, 129, 0.2); }
            50% { box-shadow: 0 4px 25px rgba(16, 185, 129, 0.5); }
        }
        @keyframes shimmer {
            0% { background-position: -200% center; }
            100% { background-position: 200% center; }
        }
        .resume-animated-col {
            opacity: 0;
            animation: resEnter 0.8s cubic-bezier(0.16, 1, 0.3, 1) forwards;
        }
        .resume-float-wrapper {
            border: 1px solid rgba(255,255,255,0.1);
            border-radius: 12px;
            padding: 6px;
            background: rgba(255,255,255,0.02);
            animation: resFloat 5s ease-in-out infinite;
            transition: all 0.4s cubic-bezier(0.16, 1, 0.3, 1);
            cursor: pointer;
            overflow: hidden;
        }
        .resume-float-wrapper:hover {
            animation-play-state: paused;
            transform: translateY(-12px) scale(1.03) !important;
            border-color: #10B981;
            box-shadow: 0 20px 40px rgba(16, 185, 129, 0.25), 0 0 0 1px rgba(16, 185, 129, 0.3);
        }
        .res-template-name {
            text-align: center;
            font-family: 'Fira Code', monospace;
            font-weight: 600;
            margin-top: 14px;
            color: #F8FAFC;
            font-size: 0.9rem;
            letter-spacing: 0.5px;
            text-transform: uppercase;
        }
        .res-ats-badge {
            background: linear-gradient(135deg, #10B981, #059669);
            background-size: 200% auto;
            animation: shimmer 3s linear infinite, atsGlow 3s ease-in-out infinite;
            color: #000;
            font-family: 'Fira Code', monospace;
            font-weight: 800;
            padding: 8px 14px;
            border-radius: 8px;
            text-align: center;
            margin-top: 10px;
            font-size: 1.05rem;
        }

        /* Mobile Swipe Carousel for Resumes */
        @media (max-width: 768px) {
            [data-testid="stHorizontalBlock"]:has(.resume-float-wrapper) {
                display: flex !important;
                flex-direction: row !important;
                width: 100% !important;
                overflow: hidden !important;
                gap: 0 !important;
            }
            [data-testid="stHorizontalBlock"]:has(.resume-float-wrapper) > [data-testid="column"] {
                min-width: 100% !important;
                width: 100% !important;
                flex: 0 0 100% !important;
                animation: mobileSwipe 12s infinite cubic-bezier(0.8, 0, 0.2, 1) !important;
                padding: 0 10px !important;
            }
            @keyframes mobileSwipe {
                0%, 20% { transform: translateX(0); }
                25%, 45% { transform: translateX(-100%); }
                50%, 70% { transform: translateX(-200%); }
                75%, 95% { transform: translateX(-300%); }
                100% { transform: translateX(0); }
            }
        }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("<h4 style='text-align:center; margin-bottom: 20px;'>ATS-Optimized Resume Formats</h4>", unsafe_allow_html=True)
    r1, r2, r3, r4 = st.columns(4)

    for idx, (col, res) in enumerate(zip([r1, r2, r3, r4], resume_data)):
        delay_enter = 0.2 * (idx + 1)
        delay_float = 1.2 * idx
        with col:
            st.markdown(f'<div class="resume-animated-col" style="animation-delay: {delay_enter}s;"><div class="resume-float-wrapper" style="animation-delay: {delay_float}s;">', unsafe_allow_html=True)
            st.image(os.path.join(ASSETS_DIR, res["file"]), use_container_width=True)
            st.markdown(f'</div><div class="res-template-name">{res["name"]}</div><div class="res-ats-badge">ATS: {res["score"]}/100</div></div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="section-title">Ready to Upgrade?</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-subtitle">Initialize your project and step into the professional tier.</div>', unsafe_allow_html=True)
    
    col_btn_1, col_btn_2, col_btn_3 = st.columns([1, 1, 1])
    with col_btn_2:
        if st.button("INITIALIZE CONNECTION", type="primary", use_container_width=True):
            st.session_state.show_form = not st.session_state.show_form
            
    if st.session_state.show_form:
        st.markdown("<br>", unsafe_allow_html=True)
        with st.form("contact_form", clear_on_submit=True):
            name = st.text_input("Name")
            email = st.text_input("Email")
            col_code, col_phone = st.columns([1, 3])
            with col_code:
                country_code = st.selectbox("Code", ["🇮🇳 +91", "🇺🇸 +1", "🌎 Other"])
            with col_phone:
                phone = st.text_input("Phone Number")
                
            service = st.selectbox("Service Required", ["Portfolio Website", "Resume Building", "Project Development"])
            message = st.text_area("Project Details")
            uploaded_file = st.file_uploader("Attach Reference File (Optional)")
            
            submit_btn = st.form_submit_button("Submit Request", type="primary")
            
            if submit_btn:
                if name and email:
                    full_phone = f"{country_code.split(' ')[1] if 'Other' not in country_code else ''} {phone}".strip()
                    final_message = message
                    if uploaded_file is not None:
                        os.makedirs(UPLOAD_DIR, exist_ok=True)
                        safe_filename = f"{int(time.time())}_{uploaded_file.name}"
                        file_path = os.path.join(UPLOAD_DIR, safe_filename)
                        with open(file_path, "wb") as f:
                            f.write(uploaded_file.getbuffer())
                        final_message += f"\n\n[Attached File: {safe_filename}]"
                        
                    try:
                        conn = sqlite3.connect(DB_PATH)
                        cursor = conn.cursor()
                        cursor.execute('''INSERT INTO client_inquiries (name, email, phone_number, service_required, requirements)
                                          VALUES (?, ?, ?, ?, ?)''', (name, email, full_phone, service, final_message))
                        conn.commit()
                        conn.close()
                        st.success(f"Request received successfully! Sending confirmation to {email}...")
                        st.balloons()
                        st.session_state.show_form = False
                        st.rerun()
                    except Exception as e:
                        st.error(f"Database Error: {e}")
                else:
                    st.error("Please provide both your name and email address.")
                    
    # Footer
    st.markdown("""
    <div class="footer">
        © 2026 Nexus Digital. All rights reserved.<br>
        Engineered for Excellence.<br><br>
        <span style="background: linear-gradient(135deg, #10B981, #059669); -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text; font-weight: 800; font-size: 1rem; letter-spacing: 2px; font-family: 'Inter', sans-serif;">NEXUS GROUP</span>
    </div>
    """, unsafe_allow_html=True)


# =========================================================
# ADMIN PANEL
# =========================================================
elif st.session_state.page == "Admin":
    st.markdown('<div class="section-title" style="text-align: left;">System Administration</div>', unsafe_allow_html=True)
    st.markdown("---")
    
    def logout():
        st.session_state.admin_pwd = ""
        st.session_state.page = "Main App"

    pwd = st.text_input("Enter Admin Key", type="password", key="admin_pwd")
    if pwd == "0211Dhidhi": 
        col1, col2 = st.columns([8, 2])
        with col1:
            st.success("Access Granted")
        with col2:
            st.button("Logout", on_click=logout)
        
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT id, name, email, phone_number, service_required, requirements, submitted_at, status FROM client_inquiries ORDER BY submitted_at DESC")
        rows = cursor.fetchall()
        conn.close()
        
        st.write(f"Total Leads Found: {len(rows)}")
        
        if len(rows) > 0:
            import pandas as pd
            df = pd.DataFrame(rows, columns=["ID", "Name", "Email", "Phone", "Service", "Requirements", "Submitted At", "Status"])
            st.dataframe(df, use_container_width=True)
            
            # Export logic
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Client Leads"
            headers = ["ID", "Name", "Email", "Phone Number", "Service Required", "Requirements", "Submitted At", "Status"]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
            for row_idx, row in enumerate(rows, 2):
                for col_idx, val in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            st.download_button(
                label="📥 Export Leads to Excel",
                data=buffer,
                file_name=f"client_leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    elif pwd:
        st.error("Access Denied: Invalid Key")
